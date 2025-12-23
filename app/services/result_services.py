from app.database.session import SessionLocal
from app.models.result import Result
from app.models.exam import Exam
from app.models.course import Course
from app.models.user import User
import csv
import io
import openpyxl
from datetime import datetime

def get_user_exam_history(user_id):
    db = SessionLocal()
    results = db.query(Result).filter_by(user_id=user_id).all()
    history = []
    for r in results:
        exam = db.query(Exam).filter_by(id=r.exam_id).first()
        course = db.query(Course).filter_by(id=exam.course_id).first() if exam else None
        history.append({
            "course": course.name if course else "Unknown",
            "exam": exam.name if exam else "Unknown",
            "score": r.score,
            "percentage": r.percentage,
            "completed_at": r.completed_at
        })
    db.close()
    return history

def get_exam_analytics():
    db = SessionLocal()
    results = db.query(Result).all()
    total_results = len(results)
    avg_score = sum(r.score for r in results) / total_results if total_results else 0
    avg_percentage = sum(r.percentage for r in results) / total_results if total_results else 0

    analytics = {
        "total_results": total_results,
        "avg_score": avg_score,
        "avg_percentage": avg_percentage
    }

    db.close()
    return analytics

def get_all_results():
    """Get all exam results with user and exam details"""
    db = SessionLocal()
    results = db.query(Result).join(User).join(Exam).all()
    detailed_results = []
    for r in results:
        course = db.query(Course).filter_by(id=r.exam.course_id).first() if r.exam else None
        detailed_results.append({
            "user_id": r.user.id,
            "username": r.user.username,
            "telegram_id": r.user.telegram_id,
            "course_name": course.name if course else "Unknown",
            "exam_name": r.exam.name,
            "score": r.score,
            "percentage": r.percentage,
            "completed_at": r.completed_at.strftime("%Y-%m-%d %H:%M:%S") if r.completed_at else None,
            "status": "Passed" if r.percentage >= 70 else "Failed"
        })
    db.close()
    return detailed_results

def export_results_csv():
    """Export all results as CSV string"""
    results = get_all_results()
    if not results:
        return None

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=results[0].keys())
    writer.writeheader()
    writer.writerows(results)
    return output.getvalue()

def export_results_excel():
    """Export all results as Excel bytes"""
    results = get_all_results()
    if not results:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"

    # Write headers
    headers = list(results[0].keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, result in enumerate(results, 2):
        for col_num, key in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=result[key])

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output